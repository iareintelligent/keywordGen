from openpyxl import load_workbook
import numpy as np
import re
import string

wb = load_workbook("tekla-2018-mar-easter.xlsx")
ws = wb.active
# seedCols = np.asarray(ws['A:C'])
col1 = np.asarray(ws['A'])
col2 = np.asarray(ws['B'])
col3 = np.asarray(ws['C'])
col4 = ws['D']
broadKeywords = []
phraseKeywords = []
exactKeywords = []


def broadMatch(seed1, seed2, seed3):
    # check to see if first word already has broad match modifier and ignore it to keep it
    # store that str index for start loc for keyword building
    i = 0
    if seed1[0] == "+":
        i = 1
    keyword = "+" + seed1[i:] + " +" + seed2 + " +" + seed3
    broadKeywords.append(keyword)


def phraseMatch(seed1, seed2, seed3):
    pattern = re.compile('[^a-zA-Z0-9\s]')
    keyword = '"' + re.sub(pattern, "", seed1) + " " + re.sub(pattern,
                                                              "", seed2) + " " + re.sub(pattern, "", seed3) + '"'
    phraseKeywords.append(keyword)


def exactMatch(seed1, seed2, seed3):
    pattern = re.compile('[^a-zA-Z0-9\s]')
    keyword = '[' + re.sub(pattern, "", seed1) + " " + re.sub(pattern,
                                                              "", seed2) + " " + re.sub(pattern, "", seed3) + ']'
    exactKeywords.append(keyword)


for x in range(1, col1.size):
    if col1[x].value != None:
        for y in range(1, col2.size):
            if col2[y].value != None:
                for z in range(1, col3.size):
                    if col3[z].value != None:
                        broadMatch(col1[x].value, col2[y].value, col3[z].value)
                        exactMatch(col1[x].value, col2[y].value, col3[z].value)
                        phraseMatch(col1[x].value,
                                    col2[y].value, col3[z].value)
                # broadMatch(cell1.value, cell2.value, cell3.value);
                # phraseMatch() will be called by exactMatch()

for x in range(0, len(broadKeywords)):
    y = x+2
    # print(y, broadKeywords[x])
    mycell = ws.cell(y, 4)
    mycell.value = broadKeywords[x]

print('doneski?')
# for i = 0; i < broadKeywords.length; i++
# ws.col4[i] = broadKeywords[i]
# ws.col4[2*i] = exactKeywords[i]
# ws.col4[3*i] = phraseKeywords[i]

# exactMatch(seed1, seed2, seed3)
# strip out all non alphanumeric characters (also ignores broad match mods)
# keyword = "[" +
#     seed1.alphaNumeric + seed2.alphaNumeric + seed3.alphaNumeric +
#     "]"
# exactKeywords.push(keyword)
# phraseMatch(keyword)

# phraseMatch(exactKeyword)
#     keyword = keyword.replace('[', '"').replace(']', '"')
#     phraseKeywords.push(keyword)
